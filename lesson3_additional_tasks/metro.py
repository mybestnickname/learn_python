from openpyxl import load_workbook
from decimal import Decimal


def dict_creater(file_name):
    """
    Функция создаёт словарь с данными о метро
    {'название странции':[геоданные, статус ремонта эскалатора]}
    """
    workbook = load_workbook(filename=file_name)
    sheets = workbook.get_sheet_names()
    # берём первый лист
    ws = workbook.get_sheet_by_name(sheets[0])
    # почему не через ws.iter_rows(row_offset=1)?
    metro_dict = {}
    #   {
    #    'имя станции':
    #       [[{'longitude':decimal_value,'latitude':decimal_value},...],
    #       ремонт эскалатора?],
    #   ...
    #    }
    for row in range(2, ws.max_row):
        # к чему эта проверка если и так итерируемся до последней
        # заполненной(даже не до активной) ячейки?
        if not ws.cell(row=row, column=2).value:
            break
        # вынимаем интересующие нас значения
        subway_stop_name = ws.cell(row=row, column=5).value
        escalators = ws.cell(row=row, column=12).value
        geo_dict = {
            'longitude': Decimal(ws.cell(row=row, column=3).value),
            'latitude': Decimal(ws.cell(row=row, column=4).value)
        }
        # заносим из в словарь
        if not metro_dict.get(subway_stop_name, False):
            metro_dict[subway_stop_name] = [[geo_dict], escalators]
        else:
            # добавляем координаты очередного выхода
            metro_dict[subway_stop_name][0].append(geo_dict)
            # если ремонт, то добовляем, если нет, то не трогаем
            if escalators:
                metro_dict[subway_stop_name][1] = escalators
    return metro_dict


def escalator_checker(metro_dict):
    """
    Функция проходится по словарю и выводит имена станций,
    на которых эскалатор находится в ремонте
    """
    print('Ремонт эскалатора на станциях:')
    for station_name, value in metro_dict.items():
        if value[1]:
            print(station_name)


if __name__ == '__main__':
    file_name = 'metro.xlsx'
    metro_dict = dict_creater(file_name)
    escalator_checker(metro_dict)
else:
    pass
